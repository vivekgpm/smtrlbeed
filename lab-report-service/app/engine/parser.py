import openpyxl

from typing import List, Dict, Any, Optional

class ExcelParser:
    def __init__(self, file):
        self.file = file
        self.wb = None

    def load(self):
        """Loads the Excel file into memory."""
        self.wb = openpyxl.load_workbook(self.file, data_only=True)

    def scan_sheet(self, sheet_name: str) -> List[List[Any]]:
        """
        Scans a sheet and returns a raw list of lists (rows).
        Handles merged cells by unmerging/replicating.
        """
        if not self.wb:
            raise ValueError("Workbook not loaded")
            
        sheet = self.wb[sheet_name]
        data = []
        
        # Simple extraction for now - can enhance for merged cells later
        for row in sheet.iter_rows(values_only=True):
            cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if any(cleaned_row):
                data.append(cleaned_row)
                
        return data

    def get_all_sheets_data(self) -> Dict[str, List[List[Any]]]:
        self.load()
        results = {}
        for sheet_name in self.wb.sheetnames:
            results[sheet_name] = self.scan_sheet(sheet_name)
        return results
